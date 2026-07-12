import calendar
import csv
import json
import re
import shutil
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

from 数据库 import ROOT, DATA_DIR, get_conn


COUNTRIES = {"ES": "西班牙", "DE": "德国", "US": "美国", "GB": "英国", "UK": "英国", "FR": "法国", "IT": "意大利"}
COUNTRY_CODES = {value: key for key, value in COUNTRIES.items()}
SUPPORTED_COUNTRIES = ["德国", "西班牙", "美国"]
SALES_FIELDS = ["月销售额(€)", "月销售额($)", "月销售额(£)", "月销售额", "Monthly Sales", "Monthly Revenue"]
PRICE_FIELDS = ["价格(€)", "价格($)", "价格(£)", "价格", "Price"]


def currency_symbol(country):
    return {"美国": "$", "英国": "£"}.get(country, "€")


def month_key(year, month):
    return f"{int(year):04d}-{int(month):02d}"


def month_int(value):
    year, month = str(value).split("-")
    return int(year) * 100 + int(month)


def normalize_month(value):
    text = str(value or "").strip()
    match = re.search(r"(\d{4})[-/年.](\d{1,2})", text)
    if not match:
        return ""
    month = int(match.group(2))
    if month < 1 or month > 12:
        return ""
    return month_key(match.group(1), month)


def month_count(start_month, end_month):
    start_month = normalize_month(start_month)
    end_month = normalize_month(end_month)
    if not start_month or not end_month:
        return 0
    start_year, start_m = [int(x) for x in start_month.split("-")]
    end_year, end_m = [int(x) for x in end_month.split("-")]
    return (end_year - start_year) * 12 + (end_m - start_m) + 1


def iter_months(start_month, end_month):
    start_month = normalize_month(start_month)
    end_month = normalize_month(end_month)
    if not start_month or not end_month:
        return []
    start_year, start_m = [int(x) for x in start_month.split("-")]
    end_year, end_m = [int(x) for x in end_month.split("-")]
    months = []
    year, month = start_year, start_m
    while (year, month) <= (end_year, end_m):
        months.append(month_key(year, month))
        month += 1
        if month > 12:
            year += 1
            month = 1
    return months


def resolve_analysis_range(conn, start_month=None, end_month=None):
    start_month = normalize_month(start_month)
    end_month = normalize_month(end_month)
    if not start_month:
        row = conn.execute("SELECT year, month FROM sales_rows ORDER BY year, month LIMIT 1").fetchone()
        if row:
            start_month = month_key(row["year"], row["month"])
    if not end_month:
        row = conn.execute("SELECT year, month FROM sales_rows ORDER BY year DESC, month DESC LIMIT 1").fetchone()
        if row:
            end_month = month_key(row["year"], row["month"])
    if start_month and end_month and month_int(start_month) > month_int(end_month):
        start_month, end_month = end_month, start_month
    return start_month, end_month

SOURCE_DATA_ROOTS = [
    ROOT / "卖家精灵数据表",
]
EXTRA_SOURCE_DIRS = []
CATEGORY_TREE_PATH = ROOT / "知识库" / "庭院细分类目汇总.xlsx"
LOCAL_CATEGORY_TREE_PATH = ROOT / "知识库" / "类目树-多语言对照.csv"

MONTH_TEXT = {
    "春季": [3, 4, 5],
    "春天": [3, 4, 5],
    "夏季": [6, 7, 8],
    "夏天": [6, 7, 8],
    "秋季": [9, 10, 11],
    "秋天": [9, 10, 11],
    "冬季": [12, 1, 2],
    "冬天": [12, 1, 2],
}
WEEKDAY_TEXT = {"一": 0, "二": 1, "三": 2, "四": 3, "五": 4, "六": 5, "日": 6, "天": 6}
CATEGORY_SYNONYMS = {
    "电钢琴": ["电子键盘", "Electronic Keyboards", "digital piano"],
    "电子琴": ["电子键盘", "Electronic Keyboards"],
    "架子鼓": ["架子鼓套装", "鼓与打击乐器"],
    "贝斯": ["吉他贝斯音箱", "吉他贝斯配件"],
    "小提琴": ["弦乐器", "violin"],
}


CATEGORY_ZH = {
    "Motosierras": "链锯",
    "Piscinas desmontables": "可拆装泳池",
    "Limpiafondos automáticos para piscinas": "自动泳池清洁机",
    "Cloro para piscinas y bañeras de hidromasaje": "泳池和热水浴缸氯剂",
    "Tumbonas": "户外躺椅",
    "Filtros para piscinas": "泳池过滤器",
    "Generadores para exteriores": "户外发电机",
    "Pulverizadores de jardinería": "园艺喷雾器",
    "Fertilizantes multiusos": "通用肥料",
    "Tierras para multiusos": "通用种植土",
    "Fundas para piscinas": "泳池罩",
    "Rasendünger": "草坪肥料",
    "Schneckenabwehr": "防蜗牛用品",
    "Unkrautvernichter": "除草剂",
    "Chlor": "泳池氯剂",
    "Fliegenabwehr": "防苍蝇用品",
    "Fallen zur Schädlingsbekämpfung": "害虫诱捕器",
    "Bienenabwehr, Wespenabwehr & Hornissenabwehr": "蜜蜂/黄蜂/马蜂防护用品",
}

PATH_SEGMENT_ZH = {
    "Jardín": "庭院",
    "Garten": "庭院",
    "Jardinería": "园艺",
    "Control de plagas y protección de plantas": "害虫控制和植物保护",
    "Control de insectos": "昆虫防治",
    "Piscinas, bañeras de hidromasaje y suministros": "泳池、热水浴缸及用品",
    "Barbacoa y comedor al aire libre": "烧烤和户外餐饮",
    "Herramientas de jardinería y equipos de riego": "园艺工具和灌溉设备",
    "Gartenarbeit": "园艺",
    "Rasenpflege": "草坪护理",
}

CHAT_CONTEXT = {"country": None, "updated_at": None}
CONTEXT_TTL_SECONDS = 300
TRANSLATION_CACHE = None
CATEGORY_TREE_CACHE = None


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def reset_translation_cache():
    global TRANSLATION_CACHE
    TRANSLATION_CACHE = None


def load_translation_map():
    global TRANSLATION_CACHE
    if TRANSLATION_CACHE is not None:
        return TRANSLATION_CACHE
    data = {}
    try:
        with get_conn() as conn:
            rows = conn.execute("SELECT original_name, zh_name FROM translations WHERE TRIM(zh_name) <> ''").fetchall()
            data = {row["original_name"]: row["zh_name"] for row in rows}
    except Exception:
        data = {}
    TRANSLATION_CACHE = data
    return data


def reset_category_tree_cache():
    global CATEGORY_TREE_CACHE
    CATEGORY_TREE_CACHE = None


def clean_text(value):
    return re.sub(r"[\s\-_&/(),.·，。:：]+", "", str(value or "").lower())


def text_tokens(value):
    raw_tokens = re.findall(r"[A-Za-zÀ-ÿÄÖÜäöüßñÑ]+|[\u4e00-\u9fff]{2,}", str(value or "").lower())
    tokens = set()
    for token in raw_tokens:
        if re.fullmatch(r"[A-Za-zÀ-ÿÄÖÜäöüßñÑ]+", token):
            for suffix in ("es", "s"):
                if token.endswith(suffix) and len(token) > len(suffix) + 3:
                    token = token[:-len(suffix)]
                    break
        if len(token) >= 2:
            tokens.add(token)
    return tokens


def category_aliases(row):
    aliases = []
    for key in ("level1_en", "level1_es", "level1_de", "level1_zh", "level2_en", "level2_es", "level2_de", "level2_zh", "level3_en", "level3_es", "level3_de", "level3_zh"):
        value = str(row.get(key, "") or "").strip()
        if value and value not in aliases:
            aliases.append(value)
    return aliases


def split_foreign_segment(parts, expected=3):
    values = [str(part or "").strip() for part in parts if str(part or "").strip()]
    if len(values) <= expected:
        return values + [""] * (expected - len(values))
    base = len(values) // expected
    extra = len(values) % expected
    result, pos = [], 0
    for index in range(expected):
        take = base + (1 if index < extra else 0)
        result.append(", ".join(values[pos:pos + take]).strip())
        pos += take
    return result


def read_multilingual_category_tree(path):
    text = Path(path).read_text(encoding="utf-8-sig")
    records = []
    for line in text.splitlines()[1:]:
        if not line.strip():
            continue
        parts = [part.strip() for part in line.split(",")]
        zh_positions = [i for i, part in enumerate(parts) if re.search(r"[\u4e00-\u9fff]", part)]
        if len(zh_positions) < 2:
            continue
        z1, z2 = zh_positions[0], zh_positions[1]
        z3 = zh_positions[2] if len(zh_positions) > 2 else len(parts)
        l1_en, l1_es, l1_de = split_foreign_segment(parts[:z1])
        l2_en, l2_es, l2_de = split_foreign_segment(parts[z1 + 1:z2])
        l3_en, l3_es, l3_de = split_foreign_segment(parts[z2 + 1:z3])
        level2_zh = parts[z2].strip()
        level3_zh = (parts[z3].strip() if z3 < len(parts) else "") or level2_zh
        records.append({
            "level1_en": l1_en,
            "level1_es": l1_es,
            "level1_de": l1_de,
            "level1_zh": parts[z1].strip(),
            "level2_en": l2_en,
            "level2_es": l2_es,
            "level2_de": l2_de,
            "level2_zh": level2_zh,
            "level3_en": l3_en or l2_en,
            "level3_es": l3_es or l2_es,
            "level3_de": l3_de or l2_de,
            "level3_zh": level3_zh,
        })
    return records


def load_category_tree_rows():
    global CATEGORY_TREE_CACHE
    if CATEGORY_TREE_CACHE is not None:
        return CATEGORY_TREE_CACHE
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute("SELECT * FROM category_tree")]
    CATEGORY_TREE_CACHE = rows
    return rows


def category_score(row, path_parts, major, minor, use_similarity=True):
    candidates = [minor, major] + path_parts
    clean_candidates = [clean_text(c) for c in candidates if str(c or "").strip()]
    level1_aliases = [row.get("level1_en"), row.get("level1_es"), row.get("level1_de"), row.get("level1_zh")]
    level2_aliases = [row.get("level2_en"), row.get("level2_es"), row.get("level2_de"), row.get("level2_zh")]
    level3_aliases = [row.get("level3_en"), row.get("level3_es"), row.get("level3_de"), row.get("level3_zh")]
    score, method, level3_score = 0, "", 0
    for alias in level3_aliases:
        alias_clean = clean_text(alias)
        if not alias_clean:
            continue
        alias_tokens = {token for token in text_tokens(alias) if len(token) >= 4}
        if alias_clean in clean_candidates:
            level3_score = max(level3_score, 100)
            method = "精确匹配"
        elif any(alias_clean in c or c in alias_clean for c in clean_candidates if len(c) >= 3):
            level3_score = max(level3_score, 72)
            method = method or "模糊匹配"
        elif alias_tokens and any(alias_tokens & {token for token in text_tokens(candidate) if len(token) >= 4} for candidate in candidates):
            level3_score = max(level3_score, 58)
            method = method or "模糊匹配"
        elif use_similarity:
            best = max([SequenceMatcher(None, alias_clean, c).ratio() for c in clean_candidates] or [0])
            if best >= 0.72:
                level3_score = max(level3_score, int(48 * best))
                method = method or "相近意思匹配"
    if not level3_score:
        return 0, ""
    score += level3_score
    for aliases, weight in ((level2_aliases, 18), (level1_aliases, 12)):
        for alias in aliases:
            alias_clean = clean_text(alias)
            if not alias_clean:
                continue
            if alias_clean in clean_candidates:
                score += weight
            elif any(alias_clean in c or c in alias_clean for c in clean_candidates if len(c) >= 3):
                score += max(8, weight - 18)
            elif use_similarity:
                best = max([SequenceMatcher(None, alias_clean, c).ratio() for c in clean_candidates] or [0])
                if best >= 0.72:
                    score += max(4, int((weight - 30) * best))
    return score, method


def match_category(category_path, major, minor, use_similarity=True):
    rows = load_category_tree_rows()
    path_parts = [p.strip() for p in str(category_path or "").split(":") if p.strip()]
    if not rows:
        return {
            "l1": path_zh(path_parts[0]) if path_parts else category_zh(major),
            "l2": path_zh(path_parts[1]) if len(path_parts) > 1 else category_zh(major),
            "l3": category_zh(minor) or category_zh(major),
            "method": "未加载类目树",
        }
    best_row, best_score, best_method = None, 0, ""
    for row in rows:
        score, method = category_score(row, path_parts, major, minor, use_similarity=use_similarity)
        if score > best_score:
            best_row, best_score, best_method = row, score, method
    if best_row and best_score >= 45:
        return {
            "l1": best_row.get("level1_zh") or best_row.get("level1_en") or "",
            "l2": best_row.get("level2_zh") or best_row.get("level2_en") or "",
            "l3": best_row.get("level3_zh") or best_row.get("level2_zh") or best_row.get("level3_en") or best_row.get("level2_en") or category_zh(minor),
            "method": best_method or "相近意思匹配",
        }
    fallback_l2 = category_zh(path_parts[1]) if len(path_parts) > 1 else category_zh(major)
    return {
        "l1": category_zh(path_parts[0]) if path_parts else category_zh(major),
        "l2": fallback_l2,
        "l3": category_zh(minor) or fallback_l2,
        "method": "未匹配，沿用原始类目",
    }


def match_category_key(category_key):
    return match_category("", "", category_key)


def category_zh(name):
    lines = str(name or "").splitlines()
    clean = lines[0].strip() if lines else ""
    if not clean:
        return "未识别类目"
    translations = load_translation_map()
    if clean in translations:
        return translations[clean]
    if clean in CATEGORY_ZH:
        return CATEGORY_ZH[clean]
    for key, value in CATEGORY_ZH.items():
        if key.lower() in clean.lower():
            return value
    return clean


def path_zh(path):
    parts = [p.strip() for p in str(path or "").split(":") if p.strip()]
    translations = load_translation_map()
    return " > ".join(PATH_SEGMENT_ZH.get(p, translations.get(p, CATEGORY_ZH.get(p, p))) for p in parts)


def to_float(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else 0.0


def parse_file_meta(path):
    match = re.search(r"([A-Z]{2})-(\d{4})\.(\d{2})", Path(path).name, re.I)
    if not match:
        ym = re.search(r"(\d{4})\.(\d{2})", Path(path).name)
        if ym:
            return None, int(ym.group(1)), int(ym.group(2))
        return None, None, None
    return COUNTRIES.get(match.group(1).upper()), int(match.group(2)), int(match.group(3))


def parse_year_month(path):
    _, year, month = parse_file_meta(path)
    return year, month


def normalize_country(country):
    text = str(country or "").strip()
    if not text:
        return ""
    if text.upper() in COUNTRIES:
        return COUNTRIES[text.upper()]
    if text in SUPPORTED_COUNTRIES or len(text) > 1:
        return text
    return text


def detect_country(message):
    text = str(message or "")
    if re.search(r"德国|德\b|德国站|DE\b", text, re.I):
        return "德国"
    if re.search(r"西班牙|西\b|西国|ES\b", text, re.I):
        return "西班牙"
    if re.search(r"美国|美\b|美国站|US\b|USA\b", text, re.I):
        return "美国"
    return ""


def get_context_country(message):
    detected = detect_country(message)
    now = datetime.now()
    if detected:
        CHAT_CONTEXT["country"] = detected
        CHAT_CONTEXT["updated_at"] = now
        return detected
    if CHAT_CONTEXT["country"] and CHAT_CONTEXT["updated_at"]:
        if (now - CHAT_CONTEXT["updated_at"]).total_seconds() <= CONTEXT_TTL_SECONDS:
            CHAT_CONTEXT["updated_at"] = now
            return CHAT_CONTEXT["country"]
    CHAT_CONTEXT["country"] = None
    CHAT_CONTEXT["updated_at"] = None
    return ""


def extract_terms(query):
    terms = [query.strip()]
    for raw, zh in CATEGORY_ZH.items():
        if zh and (zh in query or query in zh):
            terms.append(raw)
    fuzzy = [
        ("喂鸟器", "bird feeder"),
        ("喂食器", "feeder"),
        ("鸟", "bird"),
        ("泳池", "piscina"),
        ("泳池", "pool"),
        ("过滤", "filter"),
        ("过滤器", "filter"),
        ("躺椅", "tumbona"),
        ("躺椅", "liege"),
        ("链锯", "motosierra"),
        ("链锯", "kettensäge"),
        ("草坪", "rasen"),
        ("肥料", "dünger"),
        ("除草", "unkraut"),
        ("防虫", "insect"),
        ("除虫", "schädl"),
        ("蜗牛", "schnecken"),
        ("氯", "chlor"),
    ]
    for zh, raw in fuzzy:
        if zh in query:
            terms.append(raw)
    return [t for i, t in enumerate(terms) if t and t not in terms[:i]]


def meaningful_chinese_fragments(text):
    fragments = set()
    stop_words = {"什么", "时候", "什么时候", "卖得", "卖得好", "哪些", "有没有", "数据", "商家", "公司", "联系", "联系方式"}
    for block in re.findall(r"[\u4e00-\u9fff]{2,}", str(text or "")):
        max_len = min(6, len(block))
        for size in range(2, max_len + 1):
            for start in range(0, len(block) - size + 1):
                fragment = block[start:start + size]
                if fragment not in stop_words:
                    fragments.add(fragment)
    return fragments


def category_tree_terms(message):
    query_clean = clean_text(message)
    fragments = {clean_text(item) for item in meaningful_chinese_fragments(message)}
    for key, values in CATEGORY_SYNONYMS.items():
        if key in str(message or ""):
            fragments.update(clean_text(value) for value in values)
    terms = []
    for row in load_category_tree_rows():
        aliases = category_aliases(row)
        clean_aliases = [clean_text(alias) for alias in aliases]
        matched = any(alias and alias in query_clean for alias in clean_aliases)
        matched = matched or any(fragment and any(fragment in alias for alias in clean_aliases) for fragment in fragments)
        if matched:
            terms.extend(aliases)
            terms.extend([row.get("level1_zh"), row.get("level2_zh"), row.get("level3_zh")])
    result = []
    for term in terms:
        term = str(term or "").strip()
        if term and term not in result:
            result.append(term)
    return result[:18]


def read_table(path):
    path = Path(path)
    if path.suffix.lower() == ".csv":
        raw = path.read_bytes()
        for encoding in ("utf-8-sig", "gb18030", "utf-8"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="ignore")
        rows = list(csv.reader(text.splitlines()))
        if not rows:
            return [], []
        headers = [str(item or "").strip() for item in rows[0]]
        return headers, rows[1:]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(item or "").strip() for item in rows[0]]
    return headers, rows[1:]


def get_cell(row, idx, names, default=""):
    if isinstance(names, str):
        names = [names]
    for name in names:
        pos = idx.get(name)
        if pos is not None and pos < len(row):
            return row[pos]
    return default


def has_any_column(idx, names):
    return any(name in idx for name in names)


def detect_country_from_rows(headers, rows):
    idx = {name: i for i, name in enumerate(headers)}
    names = ["国家", "country", "Country", "COUNTRY", "站点", "市场", "market", "Market", "Marketplace"]
    countries = set()
    for row in rows[:200]:
        value = normalize_country(get_cell(row, idx, names))
        if value:
            countries.add(value)
    if len(countries) == 1:
        return countries.pop()
    if len(countries) > 1:
        raise ValueError("文件中包含多个国家，请拆分为单国家文件后再上传，避免数据混写。")
    return ""


def count_importable_rows(path):
    try:
        headers, rows = read_table(path)
        idx = {name: i for i, name in enumerate(headers)}
        if "ASIN" not in idx:
            return 0
        count = 0
        for row in rows:
            if row and row[idx["ASIN"]]:
                count += 1
        return count
    except Exception:
        return 0


def discover_best_sales_files():
    candidates = []
    for root in SOURCE_DATA_ROOTS:
        if root.exists():
            candidates.extend(sorted(root.glob("**/*.xlsx")))
    for folder in EXTRA_SOURCE_DIRS:
        if folder.exists():
            candidates.extend(sorted(folder.glob("*.xlsx")))

    best = {}
    for path in candidates:
        country, year, month = parse_file_meta(path)
        if not country or not year or not month:
            continue
        rows = count_importable_rows(path)
        key = (country, year, month)
        current = best.get(key)
        if not current or rows > current["rows"] or (rows == current["rows"] and path.stat().st_size > current["path"].stat().st_size):
            best[key] = {"path": path, "rows": rows}
    return [item["path"] for item in sorted(best.values(), key=lambda item: (parse_file_meta(item["path"])[0], parse_file_meta(item["path"])[1], parse_file_meta(item["path"])[2]))]


def import_sales_file(path, country=None, source_copy=False):
    path = Path(path)
    parsed_country, year, month = parse_file_meta(path)
    if not year or not month:
        return {"ok": False, "message": f"无法从文件名识别年月：{path.name}"}

    headers, rows = read_table(path)
    idx = {name: i for i, name in enumerate(headers)}
    input_country = normalize_country(country)
    data_country = detect_country_from_rows(headers, rows)
    country = input_country or data_country or parsed_country
    if not country:
        return {"ok": False, "message": "请先选择或输入数据归属国家。"}

    import_path = path
    if source_copy:
        upload_dir = DATA_DIR / "上传销售数据" / country
        upload_dir.mkdir(parents=True, exist_ok=True)
        import_path = upload_dir / path.name
        shutil.copy2(path, import_path)

    required = ["ASIN", "商品标题", "类目路径", "大类目", "小类目", "月销量", "BuyBox卖家"]
    missing = [name for name in required if name not in idx]
    if not has_any_column(idx, SALES_FIELDS):
        missing.append("月销售额(€/$)")
    if not has_any_column(idx, PRICE_FIELDS):
        missing.append("价格(€/$)")
    if missing:
        return {"ok": False, "message": f"{path.name} 缺少字段：{', '.join(missing)}"}

    parsed_rows = []
    total_rows = 0
    for row in rows:
        total_rows += 1
        asin = str(get_cell(row, idx, "ASIN") or "").strip()
        if not row or not asin:
            continue
        title = str(get_cell(row, idx, "商品标题") or "")
        brand = str(get_cell(row, idx, "品牌") or "")
        category_path = str(get_cell(row, idx, "类目路径") or "")
        major = str(get_cell(row, idx, "大类目") or "")
        minor_lines = str(get_cell(row, idx, "小类目") or "").splitlines()
        minor = minor_lines[0].strip() if minor_lines else "未识别类目"
        matched = match_category(category_path, major, minor, use_similarity=False)
        seller = str(get_cell(row, idx, "BuyBox卖家") or "").strip() or "未知卖家"
        parsed_rows.append(
            (
                country,
                year,
                month,
                asin,
                str(get_cell(row, idx, "SKU") or ""),
                title,
                brand,
                category_path,
                major,
                minor,
                to_float(get_cell(row, idx, SALES_FIELDS)),
                to_float(get_cell(row, idx, "月销量")),
                to_float(get_cell(row, idx, PRICE_FIELDS)),
                seller,
                str(get_cell(row, idx, "卖家所属地") or ""),
                str(get_cell(row, idx, "卖家信息") or ""),
                str(get_cell(row, idx, "卖家首页") or ""),
                str(get_cell(row, idx, "商品详情页链接") or ""),
                str(get_cell(row, idx, "商品主图") or ""),
                matched["l1"],
                matched["l2"],
                matched["l3"],
                matched["method"],
            )
        )

    note = "样本数偏少，适合阶段观察，不适合直接当全年结论。" if len(parsed_rows) <= 100 else ""
    category_keys = sorted({row[21] for row in parsed_rows if row[21]})
    with get_conn() as conn:
        existing = conn.execute(
            "SELECT id, imported_rows FROM raw_files WHERE file_path=?",
            (str(import_path),),
        ).fetchone()
        if existing and (existing["imported_rows"] or 0) >= len(parsed_rows):
            return {"ok": True, "message": f"已存在同名文件且样本数不少于当前文件：{import_path.name}", "file_id": existing["id"], "country": country, "year": year, "month": month, "categories": category_keys}
        if existing:
            conn.execute("DELETE FROM sales_rows WHERE file_id=?", (existing["id"],))
            conn.execute("DELETE FROM raw_files WHERE id=?", (existing["id"],))
        cur = conn.execute(
            """
            INSERT INTO raw_files(file_name, file_path, country, year, month, row_count, imported_rows, notes, imported_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (import_path.name, str(import_path), country, year, month, total_rows, len(parsed_rows), note, now_text()),
        )
        file_id = cur.lastrowid
        conn.executemany(
            """
            INSERT INTO sales_rows(
                country, year, month, asin, sku, title, brand, category_path, major_category,
                minor_category, sales, units, price, seller, seller_location, seller_info,
                seller_home, product_url, image_url, category_l1_zh, category_l2_zh,
                category_l3_zh, category_match_method, file_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [row + (file_id,) for row in parsed_rows],
        )
    return {"ok": True, "message": f"导入 {country} {import_path.name}：{len(parsed_rows)} 条数据", "file_id": file_id, "country": country, "year": year, "month": month, "categories": category_keys}


def normalize_existing_sales_categories():
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT COALESCE(category_path, '') AS category_path,
                   COALESCE(major_category, '') AS major_category,
                   COALESCE(minor_category, '') AS minor_category
            FROM sales_rows
            GROUP BY COALESCE(category_path, ''), COALESCE(major_category, ''), COALESCE(minor_category, '')
            """
        ).fetchall()
        updates = []
        for row in rows:
            key = (row["category_path"], row["major_category"], row["minor_category"])
            matched = match_category(*key, use_similarity=False)
            updates.append((matched["l1"], matched["l2"], matched["l3"], matched["method"], *key))
        conn.executemany(
            """
            UPDATE sales_rows
            SET category_l1_zh=?, category_l2_zh=?, category_l3_zh=?, category_match_method=?
            WHERE COALESCE(category_path, '')=? AND COALESCE(major_category, '')=? AND COALESCE(minor_category, '')=?
            """,
            updates,
        )


def load_category_tree():
    if LOCAL_CATEGORY_TREE_PATH.exists():
        source = LOCAL_CATEGORY_TREE_PATH.name
        records = [
            (
                row["level1_en"], row["level1_es"], row["level1_de"], row["level1_zh"],
                row["level2_en"], row["level2_es"], row["level2_de"], row["level2_zh"],
                row["level3_en"], row["level3_es"], row["level3_de"], row["level3_zh"],
                source,
            )
            for row in read_multilingual_category_tree(LOCAL_CATEGORY_TREE_PATH)
        ]
        if records:
            with get_conn() as conn:
                conn.execute("DELETE FROM category_tree WHERE source=?", (source,))
                conn.executemany(
                    """
                    INSERT INTO category_tree(
                        level1_en, level1_es, level1_de, level1_zh,
                        level2_en, level2_es, level2_de, level2_zh,
                        level3_en, level3_es, level3_de, level3_zh, source
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    records,
                )
            reset_category_tree_cache()
    if not CATEGORY_TREE_PATH.exists():
        return
    with get_conn() as conn:
        if conn.execute("SELECT COUNT(*) AS n FROM categories").fetchone()["n"]:
            return
    wb = openpyxl.load_workbook(CATEGORY_TREE_PATH, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = list(next(ws.iter_rows(values_only=True)))
    idx = {name: i for i, name in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] == "cate_id":
            continue
        rows.append(
            (
                str(row[idx.get("zh_cate2_name")] or ""),
                str(row[idx.get("zh_cate3_name")] or ""),
                str(row[idx.get("zh_cate4_name")] or ""),
                str(row[idx.get("cate2_name")] or ""),
                str(row[idx.get("cate3_name")] or ""),
                str(row[idx.get("cate4_name")] or ""),
            )
        )
    with get_conn() as conn:
        conn.executemany(
            "INSERT INTO categories(zh_cate2, zh_cate3, zh_cate4, en_cate2, en_cate3, en_cate4) VALUES (?, ?, ?, ?, ?, ?)",
            rows,
        )


def nth_weekday_of_month(year, month, nth, weekday):
    current = date(year, month, 1)
    offset = (weekday - current.weekday()) % 7
    target = current + timedelta(days=offset + (nth - 1) * 7)
    if target.month != month:
        raise ValueError("浮动节日规则超出当月范围")
    return target


def calculate_holiday_date(year, rule_text, is_floating=False):
    text = str(rule_text or "").strip()
    if not text:
        raise ValueError("日期规则不能为空")
    if is_floating:
        match = re.search(r"(\d{1,2})\s*月\s*第\s*([1-5一二三四五])\s*个?\s*星期\s*([一二三四五六日天])", text)
        if not match:
            raise ValueError(f"暂不支持的浮动日期规则：{text}")
        month = int(match.group(1))
        nth_text = match.group(2)
        nth = int(nth_text) if nth_text.isdigit() else {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5}[nth_text]
        weekday = WEEKDAY_TEXT[match.group(3)]
        return nth_weekday_of_month(year, month, nth, weekday)
    match = re.search(r"(\d{1,2})\s*(?:月|/|-)\s*(\d{1,2})\s*(?:日)?", text)
    if not match:
        raise ValueError(f"暂不支持的固定日期规则：{text}")
    return date(year, int(match.group(1)), int(match.group(2)))


def seed_holiday_rules():
    rows = [
        ("西班牙", "元旦", "Año Nuevo", "1月1日", 0, "新年、家庭聚会、年终假日消费"),
        ("西班牙", "三王节/主显节", "Día de Reyes", "1月6日", 0, "西班牙重要送礼日，儿童礼物需求强"),
        ("西班牙", "情人节", "San Valentín", "2月14日", 0, "鲜花、礼品、餐饮和情感消费节点"),
        ("西班牙", "父亲节", "Día del Padre", "3月19日", 0, "父亲节礼品、服装和家居消费"),
        ("西班牙", "母亲节", "Día de la Madre", "5月第1个星期日", 1, "鲜花、书籍、服装、美妆和礼品消费"),
        ("西班牙", "劳动节", "Día del Trabajador", "5月1日", 0, "长周末、出行、庭院户外活动"),
        ("西班牙", "圣母升天日", "Asunción de la Virgen", "8月15日", 0, "夏季长假、出游、庭院和户外用品"),
        ("西班牙", "西班牙国庆节", "Fiesta Nacional de España", "10月12日", 0, "假期出行、家庭聚会"),
        ("西班牙", "万圣节/诸圣节", "Todos los Santos", "11月1日", 0, "家庭纪念、鲜花、季节装饰"),
        ("西班牙", "黑色星期五", "Black Friday", "11月第4个星期五", 1, "线上促销，服装、配饰、礼品、家居明显增长"),
        ("西班牙", "宪法日", "Día de la Constitución", "12月6日", 0, "年末假期消费、礼品和出行"),
        ("西班牙", "圣母无染原罪节", "Inmaculada Concepción", "12月8日", 0, "年末假期消费、礼品和出行"),
        ("西班牙", "圣诞节", "Navidad", "12月25日", 0, "全年重要购物季，礼品、聚会、装饰"),
        ("德国", "元旦", "Neujahr", "1月1日", 0, "新年假期，家庭聚会和年初采购"),
        ("德国", "劳动节", "Tag der Arbeit", "5月1日", 0, "春夏户外活动、园艺和家庭维修需求"),
        ("德国", "德国统一日", "Tag der Deutschen Einheit", "10月3日", 0, "全国性假日，户外用品促销"),
        ("德国", "黑色星期五", "Black Friday", "11月第4个星期五", 1, "年末促销窗口，礼品、家居、电子和服饰需求增加"),
        ("德国", "圣诞节", "Weihnachten", "12月25日", 0, "德国全年重要购物季，礼品、家居、装饰消费"),
        ("美国", "元旦", "New Year's Day", "1月1日", 0, "年初采购、家庭聚会和折扣消费"),
        ("美国", "独立日", "Independence Day", "7月4日", 0, "户外烧烤、庭院用品、派对装饰消费"),
        ("美国", "万圣节", "Halloween", "10月31日", 0, "装饰、服装、糖果、派对用品消费"),
        ("美国", "感恩节", "Thanksgiving", "11月第4个星期四", 1, "黑五前奏，庭院、厨房用品和家庭聚会商品热卖"),
        ("美国", "黑色星期五", "Black Friday", "11月第4个星期五", 1, "全年核心促销节点，礼品、家居、电子、厨房用品需求集中"),
        ("美国", "圣诞节", "Christmas", "12月25日", 0, "礼品、装饰、聚会和家庭消费高峰"),
    ]
    with get_conn() as conn:
        for country, name_cn, name_local, rule_text, is_floating, note in rows:
            conn.execute(
                """
                INSERT INTO holiday_rules(country, name_cn, name_local, rule_text, is_floating, consumer_note, kind, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, '节日日历', ?, ?)
                ON CONFLICT(country, name_cn) DO NOTHING
                """,
                (country, name_cn, name_local, rule_text, is_floating, note, now_text(), now_text()),
            )


def refresh_holidays_for_years(years=None):
    today = date.today()
    years = sorted(set(int(y) for y in (years or [today.year - 1, today.year, today.year + 1])))
    with get_conn() as conn:
        rules = [dict(r) for r in conn.execute("SELECT * FROM holiday_rules ORDER BY country, name_cn")]
        for year in years:
            conn.execute(
                "DELETE FROM holidays WHERE year=? AND (generated_from='rule' OR rule_id IS NULL OR COALESCE(generated_from, '')='')",
                (year,),
            )
            for rule in rules:
                try:
                    target = calculate_holiday_date(year, rule["rule_text"], bool(rule["is_floating"]))
                except ValueError:
                    continue
                conn.execute(
                    """
                    INSERT INTO holidays(rule_id, country, year, start_date, end_date, name_cn, name_local, kind, consumer_note, prep_days, color, generated_from)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 120, 'blue', 'rule')
                    """,
                    (
                        rule["id"], rule["country"], year, target.isoformat(), target.isoformat(),
                        rule["name_cn"], rule["name_local"], rule["kind"], rule["consumer_note"],
                    ),
                )


def seed_holidays():
    seed_holiday_rules()
    today = date.today()
    refresh_holidays_for_years([today.year - 1, today.year, today.year + 1])


def seed_default_wisdom():
    rows = [
        ("西班牙", "三王节/主显节", "三王节是西班牙重要送礼日，儿童礼物、家庭聚会和节后折扣相关消费明显。", "三王节,主显节,送礼"),
        ("西班牙", "圣周", "圣周常伴随长假和出行，户外、园艺、防晒和旅行相关商品可提前观察。", "圣周,复活节,出行"),
        ("德国", "复活节", "德国复活节前后进入春季园艺和户外活动窗口，草坪护理、肥料、除虫和庭院维护商品值得提前备货。", "德国,复活节,园艺"),
        ("德国", "圣诞节", "德国圣诞季礼品、家居、装饰和家庭消费集中，黑五到圣诞前是重要促销窗口。", "德国,圣诞,礼品"),
    ]
    with get_conn() as conn:
        for country, title, content, keywords in rows:
            exists = conn.execute("SELECT id FROM wisdom WHERE country=? AND title=?", (country, title)).fetchone()
            if not exists:
                conn.execute(
                    "INSERT INTO wisdom(country, category_key, title, content, keywords, created_at) VALUES (?, '', ?, ?, ?, ?)",
                    (country, title, content, keywords, now_text()),
                )


def seed_initial_data():
    load_category_tree()
    seed_holidays()
    seed_default_wisdom()
    sync_all_wisdom_analysis()
    with get_conn() as conn:
        sales_count = conn.execute("SELECT COUNT(*) AS n FROM sales_rows").fetchone()["n"]
        analysis_count = conn.execute("SELECT COUNT(*) AS n FROM analysis_conclusions").fetchone()["n"]
    if sales_count and not analysis_count:
        normalize_existing_sales_categories()
        rebuild_analysis()


def rebuild_analysis(start_month=None, end_month=None, country=None):
    with get_conn() as conn:
        start_month, end_month = resolve_analysis_range(conn, start_month, end_month)
        if not start_month or not end_month:
            return {"ok": False, "message": "暂无可分析的销售数据"}
        start_int, end_int = month_int(start_month), month_int(end_month)
        if country:
            conn.execute(
                """
                DELETE FROM analysis_conclusions
                WHERE COALESCE(source, '销售数据') <> '运营智慧'
                  AND country=? AND time_range_start=? AND time_range_end=?
                """,
                (country, start_month, end_month),
            )
        else:
            conn.execute(
                """
                DELETE FROM analysis_conclusions
                WHERE COALESCE(source, '销售数据') <> '运营智慧'
                  AND time_range_start=? AND time_range_end=?
                """,
                (start_month, end_month),
            )
        country_clause = "AND country=?" if country else ""
        params = [start_int, end_int]
        if country:
            params.append(country)
        grouped = conn.execute(
            f"""
            SELECT country, year,
                   COALESCE(NULLIF(category_l3_zh, ''), minor_category) AS category_key,
                   COALESCE(MAX(category_l1_zh), '') AS category_l1_zh,
                   COALESCE(MAX(category_l2_zh), '') AS category_l2_zh,
                   COALESCE(MAX(category_l3_zh), '') AS category_l3_zh,
                   COALESCE(MAX(category_path), '') AS category_path,
                   month,
                   COUNT(*) AS sample_rows,
                   SUM(sales) AS sales,
                   SUM(units) AS units
            FROM sales_rows
            WHERE COALESCE(NULLIF(category_l3_zh, ''), minor_category) <> ''
              AND (year * 100 + month) BETWEEN ? AND ?
              {country_clause}
            GROUP BY country, year, month, COALESCE(NULLIF(category_l3_zh, ''), minor_category)
            """,
            params,
        ).fetchall()
        bucket = defaultdict(list)
        for row in grouped:
            bucket[(row["country"], row["category_key"], row["category_path"], row["category_l1_zh"], row["category_l2_zh"], row["category_l3_zh"])].append(dict(row))

        changed_categories = []
        for (country, category_key, category_path, category_l1_zh, category_l2_zh, category_l3_zh), months in bucket.items():
            months = sorted(months, key=lambda m: (m["year"], m["month"]))
            avg_sales = sum(m["sales"] or 0 for m in months) / len(months)
            peak_months = [m for m in months if (m["sales"] or 0) > avg_sales]
            if not peak_months:
                peak_months = [max(months, key=lambda m: m["sales"] or 0)]
            peak_numbers = [m["year"] * 100 + m["month"] for m in peak_months]
            placeholders = ",".join("?" for _ in peak_numbers)
            top_rows = conn.execute(
                f"""
                SELECT seller, COALESCE(MAX(seller_location), '') AS seller_location,
                       COALESCE(MAX(seller_info), '') AS seller_info,
                       COALESCE(MAX(seller_home), '') AS seller_home,
                       SUM(sales) AS sales, SUM(units) AS units
                FROM sales_rows
                WHERE country = ?
                  AND (year * 100 + month) IN ({placeholders})
                  AND (minor_category = ? OR category_l3_zh = ?)
                GROUP BY seller
                ORDER BY sales DESC
                LIMIT 10
                """,
                (country, *peak_numbers, category_key, category_key),
            ).fetchall()
            top_sellers = [dict(r) for r in top_rows]
            total_sales = sum(s["sales"] or 0 for s in top_sellers)
            total_units = sum(s["units"] or 0 for s in top_sellers)
            avg_price = total_sales / total_units if total_units else 0
            sample_months = sorted(
                [{"year": m["year"], "month": m["month"], "key": month_key(m["year"], m["month"]), "sales": round(m["sales"] or 0, 2), "rows": m["sample_rows"]} for m in months],
                key=lambda x: x["key"],
            )
            completeness_note = build_completeness_note(conn, country, start_month, end_month, sample_months)
            conn.execute(
                """
                INSERT INTO analysis_conclusions(
                    country, year, time_range_start, time_range_end, category_key, category_path, sample_months, avg_sales,
                    peak_months, price_low, price_high, top_sellers, chart_data,
                    completeness_note, created_at, category_l1_zh, category_l2_zh, category_l3_zh,
                    source, source_ref
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '销售数据', '')
                """,
                (
                    country,
                    int(start_month[:4]),
                    start_month,
                    end_month,
                    category_key,
                    category_path,
                    json.dumps(sample_months, ensure_ascii=False),
                    avg_sales,
                    json.dumps([{"year": m["year"], "month": m["month"], "key": month_key(m["year"], m["month"]), "sales": round(m["sales"] or 0, 2)} for m in peak_months], ensure_ascii=False),
                    avg_price * 0.95,
                    avg_price * 1.05,
                    json.dumps(top_sellers, ensure_ascii=False),
                    json.dumps(sample_months, ensure_ascii=False),
                    completeness_note,
                    now_text(),
                    category_l1_zh,
                    category_l2_zh,
                    category_l3_zh or category_key,
                ),
            )
            if category_key not in changed_categories:
                changed_categories.append(category_key)
            for seller in top_sellers:
                if not seller["seller"]:
                    continue
                conn.execute(
                    """
                    INSERT INTO merchant_resources(
                        country, category_key, seller_name, seller_location, seller_info, seller_home,
                        sales, units, updated_at, category_l1_zh, category_l2_zh, category_l3_zh
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(country, category_key, seller_name)
                    DO UPDATE SET sales=excluded.sales, units=excluded.units,
                                  seller_location=excluded.seller_location,
                                  seller_info=excluded.seller_info,
                                  seller_home=excluded.seller_home,
                                  updated_at=excluded.updated_at,
                                  category_l1_zh=excluded.category_l1_zh,
                                  category_l2_zh=excluded.category_l2_zh,
                                  category_l3_zh=excluded.category_l3_zh
                    """,
                    (country, category_key, seller["seller"], seller["seller_location"], seller["seller_info"], seller["seller_home"], seller["sales"] or 0, seller["units"] or 0, now_text(), category_l1_zh, category_l2_zh, category_l3_zh or category_key),
                )
        return {"ok": True, "country": country, "time_range_start": start_month, "time_range_end": end_month, "categories": changed_categories}


def format_peak_month(item, fallback_start=""):
    year = item.get("year") or (str(fallback_start)[:4] if fallback_start else "")
    month = int(item.get("month") or 0)
    return f"{year}年{month}月" if year else f"{month}月"


def format_month_key_text(value):
    text = normalize_month(value)
    if not text:
        return str(value or "")
    year, month = text.split("-")
    return f"{int(year)}年{int(month)}月"


def build_completeness_note(conn, country, start_month, end_month, sample_months):
    expected = iter_months(start_month, end_month)
    category_months = sorted({m.get("key") or month_key(m.get("year", start_month[:4]), m["month"]) for m in sample_months})
    file_rows = conn.execute(
        """
        SELECT year, month, MAX(imported_rows) AS imported_rows
        FROM raw_files
        WHERE country = ? AND (year * 100 + month) BETWEEN ? AND ?
        GROUP BY year, month
        """,
        (country, month_int(start_month), month_int(end_month)),
    ).fetchall()
    available = {month_key(row["year"], row["month"]) for row in file_rows}
    missing = [m for m in expected if m not in available]
    small = [month_key(row["year"], row["month"]) for row in file_rows if (row["imported_rows"] or 0) <= 100]
    notes = []
    if month_count(start_month, end_month) < 3:
        notes.append("数据量较少，分析结果可能不具代表性，建议扩大时间范围")
    if category_months:
        if len(category_months) < len(expected):
            notes.append(f"目前该类目仅有 {'、'.join(format_month_key_text(m) for m in category_months)}数据，其他月份暂无该类目数据；高峰判断只基于已上传月份")
        else:
            notes.append(f"该类目已有 {format_month_key_text(start_month)} 至 {format_month_key_text(end_month)} 数据")
    if missing:
        notes.append(f"{country} 在所选区间尚缺 {'、'.join(format_month_key_text(m) for m in missing)}文件")
    if small:
        notes.append(f"{'、'.join(format_month_key_text(m) for m in small)}数据源样本数偏少")
    return "；".join(notes) or "样本月份相对完整"


def build_analysis_notifications(country, start_month, end_month, category_keys):
    if not country or not start_month or not end_month or not category_keys:
        return []
    notes = []
    with get_conn() as conn:
        placeholders = ",".join("?" for _ in category_keys)
        rows = conn.execute(
            f"""
            SELECT country, time_range_start, time_range_end, category_key, peak_months, price_low, price_high, completeness_note
            FROM analysis_conclusions
            WHERE country=? AND time_range_start=? AND time_range_end=? AND category_key IN ({placeholders})
            ORDER BY category_key
            """,
            (country, start_month, end_month, *category_keys),
        ).fetchall()
        for row in rows[:8]:
            peaks = "、".join(format_peak_month(item, row["time_range_start"]) for item in json.loads(row["peak_months"]))
            symbol = currency_symbol(row["country"])
            warning = " 数据量较少，建议扩大时间范围。" if "数据量较少" in (row["completeness_note"] or "") else ""
            notes.append(f"{row['country']}-{category_zh(row['category_key'])}（{row['time_range_start']} 至 {row['time_range_end']}）分析完成：高峰月 {peaks}，价格带 {symbol}{row['price_low']:.2f}-{symbol}{row['price_high']:.2f}，已写入知识库。{warning}")
    if len(category_keys) > len(notes):
        notes.append(f"{country}-{start_month} 至 {end_month}，{len(category_keys)}个类目分析完成，已写入知识库。")
    return notes


def get_holiday_wisdom(conn, country, holiday_name):
    rows = conn.execute(
        """
        SELECT title, content
        FROM wisdom
        WHERE (country = ? OR country = '') AND (title LIKE ? OR keywords LIKE ? OR content LIKE ?)
        ORDER BY country DESC, id DESC LIMIT 1
        """,
        (country, f"%{holiday_name}%", f"%{holiday_name}%", f"%{holiday_name}%"),
    ).fetchall()
    return dict(rows[0]) if rows else None


def get_calendar(year, month):
    refresh_holidays_for_years([year])
    start = date(year, month, 1)
    grid_start = start - timedelta(days=start.weekday())
    markers = get_calendar_markers(year, month)
    by_day = defaultdict(list)
    for item in markers:
        by_day[item["date"]].append(item)
    days = []
    for i in range(42):
        current = grid_start + timedelta(days=i)
        days.append({"date": current.isoformat(), "day": current.day, "in_month": current.month == month, "items": by_day[current.isoformat()][:8]})
    return {"year": year, "month": month, "days": days, "reminders": get_active_reminders()}


def get_calendar_markers(year, month):
    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])
    markers = []
    with get_conn() as conn:
        for row in conn.execute("SELECT * FROM holidays WHERE start_date <= ? AND end_date >= ?", (month_end.isoformat(), month_start.isoformat())):
            event_start = datetime.strptime(row["start_date"], "%Y-%m-%d").date()
            if event_start.month == month and event_start.year == year:
                wisdom = get_holiday_wisdom(conn, row["country"], row["name_cn"])
                detail = wisdom["content"] if wisdom else row["consumer_note"]
                markers.append({
                    "type": "holiday",
                    "color": "blue",
                    "country": row["country"],
                    "date": event_start.isoformat(),
                    "title": row["name_cn"],
                    "detail": f"{row['country']}｜{row['name_local']}｜{detail}",
                })
        for row in conn.execute("SELECT * FROM analysis_conclusions"):
            peaks = json.loads(row["peak_months"])
            for peak in peaks:
                peak_year = int(peak.get("year") or row["year"] or str(row["time_range_start"] or year)[:4])
                target = date(peak_year, int(peak["month"]), 1)
                reminder_date = target - timedelta(days=120)
                marker_date = date(reminder_date.year, reminder_date.month, 1)
                if marker_date.year == year and marker_date.month == month:
                    symbol = currency_symbol(row["country"])
                    price_text = "" if row["source"] == "运营智慧" else f"｜价格带 {symbol}{row['price_low']:.2f}-{symbol}{row['price_high']:.2f}"
                    markers.append({
                        "type": "sales",
                        "color": "orange",
                        "country": row["country"],
                        "date": marker_date.isoformat(),
                        "title": f"{category_zh(row['category_key'])}备货提醒",
                        "detail": f"{row['country'] or '全部国家'}｜{category_zh(row['category_key'])}｜目标高峰：{target.year}年{target.month}月{price_text}｜来源：{row['source'] or '销售数据'}",
                    })
    return sorted(markers, key=lambda x: (x["date"], x["color"], x["title"]))


def get_active_reminders(today=None):
    today = today or date.today()
    refresh_holidays_for_years([today.year, today.year + 1])
    reminders = []
    with get_conn() as conn:
        for row in conn.execute("SELECT * FROM holidays"):
            target = datetime.strptime(row["start_date"], "%Y-%m-%d").date()
            reminder_start = target - timedelta(days=120)
            reminder_end = reminder_start + timedelta(days=29)
            if reminder_start <= today <= reminder_end:
                reminders.append({
                    "type": "holiday",
                    "color": "blue",
                    "country": row["country"],
                    "target_date": target.isoformat(),
                    "reminder_start": reminder_start.isoformat(),
                    "reminder_end": reminder_end.isoformat(),
                    "title": f"{row['country']} {row['name_cn']}",
                    "detail": row["consumer_note"],
                })
        for row in conn.execute("SELECT * FROM analysis_conclusions"):
            for peak in json.loads(row["peak_months"]):
                peak_year = int(peak.get("year") or row["year"] or str(row["time_range_start"] or today.year)[:4])
                target = date(peak_year, int(peak["month"]), 1)
                reminder_start = target - timedelta(days=120)
                reminder_end = reminder_start + timedelta(days=29)
                if reminder_start <= today <= reminder_end:
                    symbol = currency_symbol(row["country"])
                    reminders.append({
                        "type": "sales",
                        "color": "orange",
                        "country": row["country"],
                        "target_date": target.isoformat(),
                        "reminder_start": reminder_start.isoformat(),
                        "reminder_end": reminder_end.isoformat(),
                        "title": f"{row['country'] or '全部国家'} {category_zh(row['category_key'])}备货提醒",
                        "detail": f"目标高峰：{target.year}年{target.month}月；来源：{row['source'] or '销售数据'}；建议价格带 {symbol}{row['price_low']:.2f}-{symbol}{row['price_high']:.2f}",
                    })
    # 用户要求“越近越下方”，因此显示时远目标在上、近目标在下。
    return sorted(reminders, key=lambda x: x["target_date"], reverse=True)


def decode_row(row):
    row = dict(row)
    if not row.get("time_range_start") and row.get("year"):
        row["time_range_start"] = f"{int(row['year']):04d}-01"
    if not row.get("time_range_end") and row.get("year"):
        row["time_range_end"] = f"{int(row['year']):04d}-12"
    fallback_year = int(str(row.get("time_range_start") or row.get("year") or date.today().year)[:4])
    if "sample_months" in row:
        row["sample_months"] = json.loads(row["sample_months"])
        for item in row["sample_months"]:
            item.setdefault("year", fallback_year)
            item.setdefault("key", month_key(item["year"], item["month"]))
    if "peak_months" in row:
        row["peak_months"] = json.loads(row["peak_months"])
        for item in row["peak_months"]:
            item.setdefault("year", fallback_year)
            item.setdefault("key", month_key(item["year"], item["month"]))
    if "top_sellers" in row:
        row["top_sellers"] = json.loads(row["top_sellers"])
    if "chart_data" in row:
        row["chart_data"] = json.loads(row["chart_data"])
        for item in row["chart_data"]:
            item.setdefault("year", fallback_year)
            item.setdefault("key", month_key(item["year"], item["month"]))
    row["category_name_zh"] = category_zh(row.get("category_key"))
    row["category_path_zh"] = path_zh(row.get("category_path", ""))
    row["category_root"] = row.get("category_l1_zh") or ""
    row["time_range_label"] = f"{row.get('time_range_start', '')} 至 {row.get('time_range_end', '')}"
    return row


def get_state(year=None, month=None):
    today = date.today()
    year = int(year or today.year)
    month = int(month or today.month)
    with get_conn() as conn:
        files = [dict(r) for r in conn.execute("SELECT * FROM raw_files ORDER BY country, year, month, file_name")]
        file_categories = defaultdict(set)
        file_roots = defaultdict(set)
        for row in conn.execute("SELECT file_id, COALESCE(NULLIF(category_l3_zh, ''), minor_category) AS category_key, category_l1_zh FROM sales_rows WHERE file_id IS NOT NULL AND COALESCE(NULLIF(category_l3_zh, ''), minor_category) <> '' GROUP BY file_id, category_key, category_l1_zh"):
            file_categories[row["file_id"]].add(row["category_key"])
            if row["category_l1_zh"]:
                file_roots[row["file_id"]].add(row["category_l1_zh"])
        for row in files:
            keys = sorted(file_categories.get(row["id"], set()))
            row["category_keys"] = keys
            row["category_names_zh"] = [category_zh(key) for key in keys]
            row["category_roots"] = sorted(file_roots.get(row["id"], set()))
        conclusions = [decode_row(r) for r in conn.execute("SELECT * FROM analysis_conclusions ORDER BY time_range_end DESC, time_range_start DESC, avg_sales DESC")]
        merchants = [dict(r) for r in conn.execute("SELECT * FROM merchant_resources ORDER BY sales DESC")]
        for row in merchants:
            row["category_name_zh"] = category_zh(row["category_key"])
            row["category_root"] = row.get("category_l1_zh") or ""
        wisdom = [dict(r) for r in conn.execute("SELECT * FROM wisdom ORDER BY id DESC LIMIT 300")]
        for row in wisdom:
            row["category_name_zh"] = category_zh(row["category_key"])
        holiday_rules = [dict(r) for r in conn.execute("SELECT * FROM holiday_rules ORDER BY country, name_cn")]
        tree_count = conn.execute("SELECT COUNT(*) AS n FROM category_tree").fetchone()["n"]
        category_count = tree_count or conn.execute("SELECT COUNT(*) AS n FROM categories").fetchone()["n"]
        categories = sorted(
            {r["category_key"] for r in conn.execute("SELECT DISTINCT category_key FROM analysis_conclusions WHERE category_key <> ''")} |
            {r["category_key"] for r in conn.execute("SELECT DISTINCT category_key FROM merchant_resources WHERE category_key <> ''")} |
            {r["level3_zh"] for r in conn.execute("SELECT DISTINCT level3_zh FROM category_tree WHERE level3_zh <> ''")}
        )
        category_roots = sorted({r["category_l1_zh"] for r in conn.execute("SELECT DISTINCT category_l1_zh FROM analysis_conclusions WHERE category_l1_zh <> ''")} | {r["category_l1_zh"] for r in conn.execute("SELECT DISTINCT category_l1_zh FROM merchant_resources WHERE category_l1_zh <> ''")} | {r["level1_zh"] for r in conn.execute("SELECT DISTINCT level1_zh FROM category_tree WHERE level1_zh <> ''")})
        category_root_counts = {
            row["level1_zh"]: row["n"]
            for row in conn.execute("SELECT level1_zh, COUNT(*) AS n FROM category_tree WHERE level1_zh <> '' GROUP BY level1_zh")
        }
        country_values = set(SUPPORTED_COUNTRIES)
        for table in ("sales_rows", "raw_files", "analysis_conclusions", "merchant_resources", "wisdom", "holidays", "holiday_rules"):
            country_values.update(r["country"] for r in conn.execute(f"SELECT DISTINCT country FROM {table}") if r["country"])
        countries = sorted(country_values)
        available_months = [month_key(r["year"], r["month"]) for r in conn.execute("SELECT DISTINCT year, month FROM raw_files ORDER BY year, month")]
        time_ranges = sorted(
            {
                f"{r['time_range_start']} 至 {r['time_range_end']}"
                for r in conn.execute("SELECT DISTINCT time_range_start, time_range_end FROM analysis_conclusions WHERE COALESCE(time_range_start, '')<>'' AND COALESCE(time_range_end, '')<>''")
            },
            reverse=True,
        )
    return {
        "countries": countries,
        "today": today.isoformat(),
        "calendar": get_calendar(year, month),
        "files": files,
        "analysis": conclusions,
        "merchants": merchants,
        "wisdom": wisdom,
        "holiday_rules": holiday_rules,
        "categories": [{"key": c, "name": category_zh(c)} for c in categories],
        "category_roots": category_roots,
        "available_months": available_months,
        "time_ranges": time_ranges,
        "category_root_counts": category_root_counts,
        "category_count": category_count,
        "active_country": CHAT_CONTEXT["country"],
    }


def add_wisdom(title, content, category_key="", keywords="", country=""):
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO wisdom(country, category_key, title, content, keywords, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (normalize_country(country), category_key.strip(), title.strip() or "人工补充", content.strip(), keywords.strip(), now_text(), now_text()),
        )
        wisdom_id = cur.lastrowid
    sync_wisdom_analysis(wisdom_id)


def update_wisdom(wisdom_id, data):
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE wisdom
            SET country=?, category_key=?, title=?, content=?, keywords=?, updated_at=?
            WHERE id=?
            """,
            (
                normalize_country(data.get("country", "")),
                str(data.get("category_key") or "").strip(),
                str(data.get("title") or "").strip() or "人工补充",
                str(data.get("content") or "").strip(),
                str(data.get("keywords") or "").strip(),
                now_text(),
                int(wisdom_id),
            ),
        )
    sync_wisdom_analysis(wisdom_id)


def delete_wisdom(wisdom_id):
    with get_conn() as conn:
        conn.execute("DELETE FROM wisdom WHERE id=?", (int(wisdom_id),))
        conn.execute("DELETE FROM analysis_conclusions WHERE source='运营智慧' AND source_ref=?", (f"wisdom:{int(wisdom_id)}",))


def add_holiday_rule(data):
    country = normalize_country(data.get("country"))
    name = str(data.get("name_cn") or "").strip()
    rule_text = str(data.get("rule_text") or "").strip()
    if not country or not name or not rule_text:
        raise ValueError("国家、节日名称、日期规则必填")
    is_floating = 1 if data.get("is_floating") else 0
    calculate_holiday_date(date.today().year, rule_text, bool(is_floating))
    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO holiday_rules(country, name_cn, name_local, rule_text, is_floating, consumer_note, kind, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, '节日日历', ?, ?)
            ON CONFLICT(country, name_cn)
            DO UPDATE SET name_local=excluded.name_local, rule_text=excluded.rule_text,
                          is_floating=excluded.is_floating, consumer_note=excluded.consumer_note,
                          updated_at=excluded.updated_at
            """,
            (country, name, data.get("name_local", ""), rule_text, is_floating, data.get("consumer_note", ""), now_text(), now_text()),
        )
    today = date.today()
    refresh_holidays_for_years([today.year - 1, today.year, today.year + 1])


def update_holiday_rule(rule_id, data):
    country = normalize_country(data.get("country"))
    name = str(data.get("name_cn") or "").strip()
    rule_text = str(data.get("rule_text") or "").strip()
    if not country or not name or not rule_text:
        raise ValueError("国家、节日名称、日期规则必填")
    is_floating = 1 if data.get("is_floating") else 0
    calculate_holiday_date(date.today().year, rule_text, bool(is_floating))
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE holiday_rules
            SET country=?, name_cn=?, name_local=?, rule_text=?, is_floating=?, consumer_note=?, updated_at=?
            WHERE id=?
            """,
            (country, name, data.get("name_local", ""), rule_text, is_floating, data.get("consumer_note", ""), now_text(), int(rule_id)),
        )
    today = date.today()
    refresh_holidays_for_years([today.year - 1, today.year, today.year + 1])


def delete_holiday_rule(rule_id):
    with get_conn() as conn:
        conn.execute("DELETE FROM holiday_rules WHERE id=?", (int(rule_id),))
        conn.execute("DELETE FROM holidays WHERE rule_id=?", (int(rule_id),))


def parse_months_from_text(text):
    value = str(text or "")
    months = set()
    for word, word_months in MONTH_TEXT.items():
        if word in value:
            months.update(word_months)
    for start, end in re.findall(r"(\d{1,2})\s*(?:-|—|~|至|到)\s*(\d{1,2})\s*月", value):
        start_i, end_i = int(start), int(end)
        if 1 <= start_i <= 12 and 1 <= end_i <= 12:
            if start_i <= end_i:
                months.update(range(start_i, end_i + 1))
            else:
                months.update(list(range(start_i, 13)) + list(range(1, end_i + 1)))
    for month in re.findall(r"(\d{1,2})\s*月", value):
        month_i = int(month)
        if 1 <= month_i <= 12:
            months.add(month_i)
    return sorted(months)


def category_rows_from_text(text, fallback_key=""):
    value = str(text or "")
    for key, values in CATEGORY_SYNONYMS.items():
        if key in value:
            value += " " + " ".join(values)
    clean_value = clean_text(value)
    matched = []
    for row in load_category_tree_rows():
        level3 = str(row.get("level3_zh") or "").strip()
        level2 = str(row.get("level2_zh") or "").strip()
        level1 = str(row.get("level1_zh") or "").strip()
        aliases = [row.get("level3_zh"), row.get("level3_en"), row.get("level3_es"), row.get("level3_de")]
        if any(alias and clean_text(alias) in clean_value for alias in aliases):
            matched.append(row)
        elif level3 and level3 in value:
            matched.append(row)
        elif fallback_key and fallback_key in {level2, level3}:
            matched.append(row)
    if matched:
        unique = {}
        for row in matched:
            unique[row["level3_zh"]] = row
        return list(unique.values())
    if fallback_key:
        matched_key = match_category_key(fallback_key)
        return [{
            "level1_zh": matched_key["l1"] or fallback_key,
            "level2_zh": matched_key["l2"] or fallback_key,
            "level3_zh": matched_key["l3"] or fallback_key,
        }]
    return []


def sync_wisdom_analysis(wisdom_id):
    current_year = date.today().year
    start_month = f"{current_year}-01"
    end_month = f"{current_year}-12"
    source_ref = f"wisdom:{int(wisdom_id)}"
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM wisdom WHERE id=?", (int(wisdom_id),)).fetchone()
        conn.execute("DELETE FROM analysis_conclusions WHERE source='运营智慧' AND source_ref=?", (source_ref,))
        if not row:
            return
        text = f"{row['title']}。{row['content']}"
        months = parse_months_from_text(text)
        categories = category_rows_from_text(text, row["category_key"])
        if not months or not categories:
            return
        for category in categories:
            category_key = category.get("level3_zh") or row["category_key"]
            country = normalize_country(row["country"]) or ""
            sample_months = [{"year": current_year, "month": month, "key": month_key(current_year, month), "sales": 0, "rows": 0} for month in months]
            peak_months = [{"year": current_year, "month": month, "key": month_key(current_year, month), "sales": 0} for month in months]
            conn.execute(
                """
                INSERT INTO analysis_conclusions(
                    country, year, time_range_start, time_range_end, category_key, category_path, sample_months, avg_sales,
                    peak_months, price_low, price_high, top_sellers, chart_data,
                    completeness_note, created_at, category_l1_zh, category_l2_zh, category_l3_zh,
                    source, source_ref
                )
                VALUES (?, ?, ?, ?, ?, '', ?, 0, ?, 0, 0, '[]', ?, ?, ?, ?, ?, ?, '运营智慧', ?)
                """,
                (
                    country,
                    current_year,
                    start_month,
                    end_month,
                    category_key,
                    json.dumps(sample_months, ensure_ascii=False),
                    json.dumps(peak_months, ensure_ascii=False),
                    json.dumps(sample_months, ensure_ascii=False),
                    f"来自运营智慧库：{row['title']}。这是人工规律结论，不含销售额和价格带。",
                    now_text(),
                    category.get("level1_zh") or row["category_key"],
                    category.get("level2_zh") or row["category_key"],
                    category_key,
                    source_ref,
                ),
            )


def sync_all_wisdom_analysis():
    with get_conn() as conn:
        ids = [row["id"] for row in conn.execute("SELECT id FROM wisdom")]
    for wisdom_id in ids:
        sync_wisdom_analysis(wisdom_id)


def add_merchant(data):
    country = normalize_country(data.get("country"))
    if not country:
        raise ValueError("请先选择国家")
    seller_name = str(data.get("seller_name") or "").strip()
    category_key = str(data.get("category_key") or "").strip()
    if not seller_name or not category_key:
        raise ValueError("店铺名字和主营三级类目必填")
    matched = match_category_key(category_key)
    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO merchant_resources(country, category_key, seller_name, company_name, contact, email, address, notes, source, updated_at, category_l1_zh, category_l2_zh, category_l3_zh)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, '手动新增', ?, ?, ?, ?)
            ON CONFLICT(country, category_key, seller_name)
            DO UPDATE SET company_name=excluded.company_name, contact=excluded.contact,
                          email=excluded.email, address=excluded.address, notes=excluded.notes,
                          updated_at=excluded.updated_at,
                          category_l1_zh=excluded.category_l1_zh,
                          category_l2_zh=excluded.category_l2_zh,
                          category_l3_zh=excluded.category_l3_zh
            """,
            (country, category_key, seller_name, data.get("company_name", ""), data.get("contact", ""), data.get("email", ""), data.get("address", ""), data.get("notes", ""), now_text(), matched["l1"], matched["l2"], matched["l3"]),
        )


def update_merchant(merchant_id, data):
    matched = match_category_key(data.get("category_key", ""))
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE merchant_resources
            SET seller_name=?, company_name=?, contact=?, email=?, address=?, category_key=?, notes=?, updated_at=?,
                category_l1_zh=?, category_l2_zh=?, category_l3_zh=?
            WHERE id=?
            """,
            (data.get("seller_name", ""), data.get("company_name", ""), data.get("contact", ""), data.get("email", ""), data.get("address", ""), data.get("category_key", ""), data.get("notes", ""), now_text(), matched["l1"], matched["l2"], matched["l3"], int(merchant_id)),
        )


def delete_merchant(merchant_id):
    with get_conn() as conn:
        conn.execute("DELETE FROM merchant_resources WHERE id=?", (int(merchant_id),))


def import_merchants_file(path, country=""):
    headers, rows = read_table(path)
    idx = {name: i for i, name in enumerate(headers)}
    default_country = normalize_country(country)
    added, skipped = 0, []
    with get_conn() as conn:
        for row in rows:
            row_country = normalize_country(get_cell(row, idx, ["国家", "country", "Country"])) or default_country
            seller_name = str(get_cell(row, idx, ["店铺名字", "店铺名称", "seller_name", "Seller"]) or "").strip()
            company_name = str(get_cell(row, idx, ["公司名称", "公司名字", "company_name", "Company"]) or "").strip()
            category_key = str(get_cell(row, idx, ["主营三级类目", "三级类目", "category_key", "Category"]) or "").strip()
            if not row_country or not category_key or (not seller_name and not company_name):
                skipped.append(seller_name or company_name or "空白行")
                continue
            exists = None
            if seller_name:
                exists = conn.execute(
                    "SELECT id FROM merchant_resources WHERE country=? AND seller_name=?",
                    (row_country, seller_name),
                ).fetchone()
            if not exists and company_name:
                exists = conn.execute(
                    "SELECT id FROM merchant_resources WHERE country=? AND company_name=?",
                    (row_country, company_name),
                ).fetchone()
            if exists:
                skipped.append(seller_name or company_name)
                continue
            matched = match_category_key(category_key)
            conn.execute(
                """
                INSERT INTO merchant_resources(country, category_key, seller_name, company_name, contact, email, address, notes, source, updated_at, category_l1_zh, category_l2_zh, category_l3_zh)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, '手动批量导入', ?, ?, ?, ?)
                """,
                (
                    row_country,
                    category_key,
                    seller_name or company_name,
                    company_name,
                    str(get_cell(row, idx, ["联系电话", "电话", "contact"]) or "").strip(),
                    str(get_cell(row, idx, ["邮箱", "email", "Email"]) or "").strip(),
                    str(get_cell(row, idx, ["经营地址", "地址", "address"]) or "").strip(),
                    str(get_cell(row, idx, ["备注", "notes"]) or "").strip(),
                    now_text(),
                    matched["l1"],
                    matched["l2"],
                    matched["l3"],
                ),
            )
            added += 1
    return {"ok": True, "added": added, "skipped": skipped, "message": f"新增 {added} 个商家，跳过 {len(skipped)} 个重复或不完整条目。"}


def export_translation_terms():
    export_dir = DATA_DIR / "导出"
    export_dir.mkdir(parents=True, exist_ok=True)
    target = export_dir / f"待翻译词汇-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    terms = set()
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT category_path, major_category, minor_category
            FROM sales_rows
            """
        ).fetchall()
        for row in rows:
            values = [row["major_category"], row["minor_category"]]
            values.extend(part.strip() for part in str(row["category_path"] or "").split(":"))
            for value in values:
                value = str(value or "").strip()
                if value and re.search(r"[A-Za-zÀ-ÿÄÖÜäöüßñÑ]", value):
                    terms.add(value)
        translations = {row["original_name"]: row["zh_name"] for row in conn.execute("SELECT original_name, zh_name FROM translations")}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "待翻译词汇"
    ws.append(["原始名称", "中文翻译"])
    for term in sorted(terms, key=str.lower):
        ws.append([term, translations.get(term, "")])
    wb.save(target)
    return target


def import_translation_file(path):
    headers, rows = read_table(path)
    idx = {name: i for i, name in enumerate(headers)}
    if "原始名称" not in idx or "中文翻译" not in idx:
        raise ValueError("翻译表必须包含表头：原始名称、中文翻译")
    count = 0
    with get_conn() as conn:
        for row in rows:
            original = str(get_cell(row, idx, "原始名称") or "").strip()
            zh_name = str(get_cell(row, idx, "中文翻译") or "").strip()
            if not original:
                continue
            conn.execute(
                """
                INSERT INTO translations(original_name, zh_name, updated_at)
                VALUES (?, ?, ?)
                ON CONFLICT(original_name) DO UPDATE SET zh_name=excluded.zh_name, updated_at=excluded.updated_at
                """,
                (original, zh_name, now_text()),
            )
            count += 1
    reset_translation_cache()
    return {"ok": True, "message": f"已更新 {count} 条翻译对照。"}


def search_terms(message):
    text = str(message or "").strip()
    terms = [text]
    terms.extend(re.findall(r"[\u4e00-\u9fff]{2,}|[A-Za-zÀ-ÿÄÖÜäöüßñÑ0-9]+", text))
    terms.extend(extract_terms(text))
    terms.extend(category_tree_terms(text))
    synonym_terms = {
        "泳池": ["泳池", "游泳池", "pool", "pools", "piscina", "piscinas", "schwimmbecken", "poolpflege"],
        "草坪": ["草坪", "rasen", "lawn", "césped"],
        "肥料": ["肥料", "dünger", "fertilizante", "fertilizer"],
        "喂鸟器": ["喂鸟器", "鸟食", "bird", "birds", "aves", "vögel", "comederos"],
        "过滤器": ["过滤器", "filter", "filtros", "kartuschen"],
    }
    for key, values in synonym_terms.items():
        if key in text:
            terms.extend(values)
    if detect_country(text):
        terms.append(detect_country(text))
    for raw, zh in CATEGORY_ZH.items():
        if zh and zh in text:
            terms.extend([raw, zh])
    cleaned = []
    for term in terms:
        term = str(term or "").strip()
        if len(term) >= 2 and term not in cleaned:
            cleaned.append(term)
    return cleaned[:24]


def matches_terms(row_text, terms):
    text = str(row_text or "").lower()
    return any(term.lower() in text for term in terms)


def score_terms(row_text, terms):
    text = str(row_text or "").lower()
    score = 0
    weak_terms = {"德国", "西班牙", "美国", "英国", "法国", "意大利", "商家", "公司", "联系方式", "联系", "数据", "高峰"}
    for term in terms:
        term_text = term.lower()
        if term_text not in text:
            continue
        score += 1 if term in weak_terms or term.isdigit() else max(2, min(len(term), 6))
    return score


def detect_category_scope(message):
    query_text = str(message or "")
    for key, values in CATEGORY_SYNONYMS.items():
        if key in query_text:
            query_text += " " + " ".join(values)
    query_clean = clean_text(query_text)
    if is_enumeration_query(message):
        for row in load_category_tree_rows():
            value = row.get("level1_zh")
            value_clean = clean_text(value)
            if value_clean and value_clean in query_clean:
                return {"field": "category_l1_zh", "value": value, "level": 1, "score": 100 + len(value_clean)}
    best = None
    for row in load_category_tree_rows():
        checks = [
            ("category_l3_zh", row.get("level3_zh"), 3),
            ("category_l2_zh", row.get("level2_zh"), 2),
            ("category_l1_zh", row.get("level1_zh"), 1),
        ]
        for field, value, level in checks:
            value_clean = clean_text(value)
            if value_clean and value_clean in query_clean:
                score = level * 100 + len(value_clean)
                if not best or score > best["score"]:
                    best = {"field": field, "value": value, "level": level, "score": score}
    return best


def is_enumeration_query(message):
    return bool(re.search(r"还有|除了|哪些|别的|其他|全部|所有|列表", str(message or "")))


def is_country_overview_query(message, country, category_scope):
    if not country or category_scope:
        return False
    return bool(re.search(r"哪些|全部|所有|类目|品类|商品|销售高峰|高峰|热销|备货", str(message or "")))


def excluded_category_terms(message):
    match = re.search(r"除了(.+?)(?:还有|以外|之外|，|,|吗|$)", str(message or ""))
    if not match:
        return []
    return [term for term in meaningful_chinese_fragments(match.group(1)) if len(term) >= 2]


def in_category_scope(item, scope):
    if not scope:
        return True
    return str(item.get(scope["field"], "") or "") == str(scope["value"] or "")


def excluded_by_terms(item, excludes):
    text = " ".join(str(item.get(key, "")) for key in ["category_key", "category_name_zh", "category_l1_zh", "category_l2_zh", "category_l3_zh"])
    return any(term and term in text for term in excludes)


def fuzzy_search_all(message):
    terms = search_terms(message)
    detected_country = get_context_country(message)
    category_scope = detect_category_scope(message)
    enumerate_all = is_enumeration_query(message)
    country_overview = is_country_overview_query(message, detected_country, category_scope)
    excludes = excluded_category_terms(message)
    context = {"query": message, "terms": terms, "country": detected_country, "category_scope": category_scope, "enumerate": enumerate_all, "analysis": [], "merchants": [], "wisdom": [], "raw_rows": [], "raw_files": [], "holidays": []}
    if not terms:
        return context
    with get_conn() as conn:
        analysis_hits = []
        for row in conn.execute("SELECT * FROM analysis_conclusions ORDER BY avg_sales DESC"):
            item = decode_row(row)
            if detected_country and item["country"] not in (detected_country, "", "全部国家"):
                continue
            if not in_category_scope(item, category_scope) or excluded_by_terms(item, excludes):
                continue
            text = " ".join([
                item["country"], str(item["year"]), item.get("time_range_label", ""), item["category_key"], item.get("category_name_zh", ""),
                item.get("category_path", ""), item.get("category_path_zh", ""),
                item.get("category_l1_zh", ""), item.get("category_l2_zh", ""), item.get("category_l3_zh", ""),
                item["peak_months"].__repr__(),
            ])
            score = score_terms(text, terms)
            if score >= 2 or (enumerate_all and category_scope) or country_overview:
                analysis_hits.append((score, item))
        context["analysis"] = [item for _, item in sorted(analysis_hits, key=lambda pair: pair[0], reverse=True)[:60 if enumerate_all else 12]]

        merchant_hits = []
        for row in conn.execute("SELECT * FROM merchant_resources ORDER BY sales DESC"):
            item = dict(row)
            if detected_country and item["country"] not in (detected_country, "", "全部国家"):
                continue
            if not in_category_scope(item, category_scope) or excluded_by_terms(item, excludes):
                continue
            item["category_name_zh"] = category_zh(item["category_key"])
            text = " ".join(str(item.get(key, "")) for key in [
                "country", "seller_name", "company_name", "contact", "address", "email",
                "seller_location", "seller_info", "category_key", "category_name_zh", "category_l1_zh", "category_l2_zh", "category_l3_zh", "notes",
            ])
            score = score_terms(text, terms)
            if score >= 2 or (enumerate_all and category_scope):
                merchant_hits.append((score, item))
        context["merchants"] = [item for _, item in sorted(merchant_hits, key=lambda pair: pair[0], reverse=True)[:15]]

        wisdom_hits = []
        for row in conn.execute("SELECT * FROM wisdom ORDER BY id DESC"):
            item = dict(row)
            if detected_country and item.get("country") and item["country"] not in (detected_country, "", "全部国家"):
                continue
            item["category_name_zh"] = category_zh(item["category_key"])
            if category_scope and item.get("category_key") and item.get("category_key") != category_scope["value"]:
                matched = match_category_key(item.get("category_key"))
                item.update({"category_l1_zh": matched["l1"], "category_l2_zh": matched["l2"], "category_l3_zh": matched["l3"]})
            if not in_category_scope(item, category_scope):
                continue
            text = " ".join(str(item.get(key, "")) for key in ["country", "category_key", "category_name_zh", "title", "content", "keywords"])
            score = score_terms(text, terms)
            if score >= 2:
                wisdom_hits.append((score, item))
        context["wisdom"] = [item for _, item in sorted(wisdom_hits, key=lambda pair: pair[0], reverse=True)[:10]]

        raw_hits = []
        for row in conn.execute("SELECT * FROM sales_rows ORDER BY sales DESC"):
            item = dict(row)
            if detected_country and item["country"] != detected_country:
                continue
            if not in_category_scope(item, category_scope) or excluded_by_terms(item, excludes):
                continue
            text = " ".join(str(item.get(key, "")) for key in [
                "country", "year", "month", "title", "brand", "category_path", "major_category",
                "minor_category", "category_l1_zh", "category_l2_zh", "category_l3_zh", "seller", "seller_location", "seller_info",
            ])
            score = score_terms(text, terms)
            if score >= 2 or (enumerate_all and category_scope):
                item["category_name_zh"] = item.get("category_l3_zh") or category_zh(item["minor_category"])
                raw_hits.append((score, item))
        context["raw_rows"] = [item for _, item in sorted(raw_hits, key=lambda pair: pair[0], reverse=True)[:60 if enumerate_all else 15]]

        raw_file_hits = []
        for row in conn.execute("SELECT * FROM raw_files ORDER BY imported_at DESC"):
            item = dict(row)
            if detected_country and item["country"] != detected_country:
                continue
            text = " ".join(str(item.get(key, "")) for key in ["country", "year", "month", "file_name", "notes"])
            score = score_terms(text, terms)
            if score >= 2:
                raw_file_hits.append((score, item))
        context["raw_files"] = [item for _, item in sorted(raw_file_hits, key=lambda pair: pair[0], reverse=True)[:10]]

        holiday_hits = []
        for row in conn.execute("SELECT * FROM holidays ORDER BY start_date"):
            item = dict(row)
            if detected_country and item["country"] != detected_country:
                continue
            text = " ".join(str(item.get(key, "")) for key in ["country", "year", "start_date", "name_cn", "name_local", "kind", "consumer_note"])
            score = score_terms(text, terms)
            if score >= 2:
                holiday_hits.append((score, item))
        context["holidays"] = [item for _, item in sorted(holiday_hits, key=lambda pair: pair[0], reverse=True)[:10]]
    return context


def unique_rows(rows, key):
    seen, result = set(), []
    for row in rows:
        value = row.get(key)
        if value in seen:
            continue
        seen.add(value)
        result.append(row)
    return result


def read_deepseek_config():
    text = ""
    for path in [ROOT / "deepseek.env", ROOT / "deepseek.env.rtf"]:
        if path.exists():
            text += "\n" + path.read_text(errors="ignore")
    key_match = re.search(r"API_KEY\s*[=:]\s*([A-Za-z0-9_\-.]+)", text)
    base_match = re.search(r"API_BASE_URL\s*[=:]\s*(https?://[A-Za-z0-9_\-./]+)", text)
    return {"key": key_match.group(1) if key_match else "", "base": (base_match.group(1) if base_match else "https://api.deepseek.com/v1").rstrip("/")}


def build_context_prompt(message, context):
    lines = ["【本地数据库查到的信息】"]
    for merchant in context["merchants"]:
        lines.append(f"- 店铺：{merchant['seller_name']}，公司：{merchant.get('company_name') or '待补充'}，电话：{merchant.get('contact') or '待补充'}，邮箱：{merchant.get('email') or '待补充'}，地址：{merchant.get('address') or merchant.get('seller_location') or '待补充'}，主营类目：{category_zh(merchant.get('category_key'))}")
    for item in context["analysis"]:
        peaks = "、".join(format_peak_month(p, item.get("time_range_start", "")) for p in item["peak_months"])
        symbol = currency_symbol(item["country"])
        lines.append(f"- 类目：{item['category_name_zh']}（原始：{item['category_key']}），时间范围：{item.get('time_range_label', '')}，该类目销售高峰：{peaks}，价格带：{symbol}{item['price_low']:.2f}-{symbol}{item['price_high']:.2f}，更新时间：{item['created_at']}")
    for item in context["wisdom"]:
        lines.append(f"- 运营智慧：{item['title']}：{item['content']}")
    for item in context.get("raw_rows", [])[:8]:
        symbol = currency_symbol(item["country"])
        lines.append(f"- 原始数据：{item['country']} {item['year']}年{item['month']}月，商品：{item.get('title') or '未命名'}，类目：{item.get('category_l3_zh') or category_zh(item.get('minor_category'))}，商家：{item.get('seller') or '未知'}，销售额：{symbol}{item.get('sales') or 0:.2f}，销量：{item.get('units') or 0}")
    for item in context.get("raw_files", []):
        lines.append(f"- 原始文件：{item['country']} {item['year']}年{item['month']}月，文件名：{item['file_name']}，导入行数：{item['imported_rows']}")
    for item in context.get("holidays", []):
        lines.append(f"- 日历节庆：{item['country']} {item['start_date']} {item['name_cn']} / {item['name_local']}：{item['consumer_note']}")
    if len(lines) == 1:
        lines.append("- 本地数据库中暂无匹配信息。")
    lines.append("")
    lines.append("【用户的问题】")
    lines.append(message)
    lines.append("")
    lines.append("请基于以上【本地数据库查到的信息】回答用户的问题。如果信息不足以完整回答，请主动说明缺失内容并追问用户。严禁自行编造具体销售额数字、具体高峰月份、具体价格带和商家联系方式。")
    return "\n".join(lines)


def ask_deepseek(message, context):
    cfg = read_deepseek_config()
    if not cfg["key"]:
        return ""
    payload = {
        "model": "deepseek-chat",
        "messages": [
            {"role": "system", "content": "你是跨境电商GMV诊断助手。只能使用用户提供的本地数据库信息作答，禁止编造销售额、高峰月份、价格带、商家联系方式。"},
            {"role": "user", "content": build_context_prompt(message, context)},
        ],
        "temperature": 0.2,
    }
    req = urllib.request.Request(
        cfg["base"] + "/chat/completions",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {cfg['key']}"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            return data["choices"][0]["message"]["content"].strip()
    except (urllib.error.URLError, KeyError, json.JSONDecodeError, TimeoutError):
        return ""


def context_has_results(context):
    return any(context.get(key) for key in ("analysis", "merchants", "wisdom", "raw_rows", "raw_files", "holidays"))


def priority_context(context):
    result = dict(context)
    if context.get("analysis"):
        result.update({"wisdom": [], "merchants": [], "raw_rows": [], "raw_files": [], "holidays": []})
    elif context.get("wisdom"):
        result.update({"merchants": [], "raw_rows": [], "raw_files": [], "holidays": []})
    elif context.get("merchants"):
        result.update({"raw_rows": [], "raw_files": [], "holidays": []})
    return result


def local_answer(context, message):
    parts = []
    if context["analysis"]:
        items = context["analysis"] if context.get("enumerate") else context["analysis"][:1]
        lines = []
        for item in items[:30]:
            peaks = "、".join(format_peak_month(p, item.get("time_range_start", "")) for p in item["peak_months"])
            symbol = currency_symbol(item["country"])
            price = "" if not (item.get("price_low") or item.get("price_high")) else f"，建议价格带为 {symbol}{item['price_low']:.2f}-{symbol}{item['price_high']:.2f}"
            source = "（运营智慧）" if item.get("source") == "运营智慧" else ""
            lines.append(f"- {item['country'] or '全部国家'}｜{item['category_name_zh']}{source}｜{item.get('time_range_label', '')}：高峰月 {peaks}{price}")
        if len(lines) == 1:
            parts.append("本地分析结论库显示：" + lines[0].lstrip("- "))
        else:
            parts.append("本地分析结论库匹配到这些结果：\n" + "\n".join(lines))
        notes = [item.get("completeness_note") for item in items[:5] if item.get("completeness_note")]
        if notes:
            parts.append("口径提醒：" + "；".join(dict.fromkeys(notes)) + "。")
    if context["merchants"]:
        names, missing_contact = [], []
        for merchant in context["merchants"]:
            if merchant["seller_name"] not in names:
                names.append(merchant["seller_name"])
            if not merchant.get("contact") and not merchant.get("email"):
                missing_contact.append(merchant["seller_name"])
        parts.append(f"商家资源库可先看：{'、'.join(names[:6])}。")
        if re.search(r"联系|电话|邮箱|联系方式", message) and missing_contact:
            parts.append(f"但这些商家的联系电话/邮箱尚未完整录入：{'、'.join(missing_contact[:5])}。你可以在商家资源库里手动补充。")
    if context["wisdom"]:
        parts.append(f"运营智慧库相关内容：{context['wisdom'][0]['title']} - {context['wisdom'][0]['content'][:160]}")
    if context.get("raw_rows"):
        if context.get("enumerate"):
            category_map = {}
            for row in context["raw_rows"]:
                key = row.get("category_l3_zh") or category_zh(row.get("minor_category"))
                category_map.setdefault(key, set()).add(f"{row['year']}年{row['month']}月")
            lines = [f"- {key}：有记录月份 {'、'.join(sorted(months))}" for key, months in list(category_map.items())[:30]]
            parts.append("原始数据湖匹配到这些类目：\n" + "\n".join(lines))
        else:
            row = context["raw_rows"][0]
            parts.append(f"原始数据湖匹配到：{row['country']} {row['year']}年{row['month']}月，{row.get('category_l3_zh') or category_zh(row.get('minor_category'))}下有商品“{row.get('title') or '未命名'}”。")
    if context.get("holidays"):
        item = context["holidays"][0]
        parts.append(f"日历数据匹配到：{item['country']} {item['start_date']} {item['name_cn']}，消费习惯：{item['consumer_note']}")
    if not parts:
        parts.append("本地数据库中暂无该信息，建议你手动更新知识库。")
    return "\n".join(parts)


def source_note(context):
    sources = []
    if context["analysis"]:
        latest = max((a.get("created_at") or "" for a in context["analysis"]), default="")
        sources.append(f"本地分析结论库，更新时间：{latest or '未知'}")
    if context["wisdom"]:
        latest = max((w.get("created_at") or "" for w in context["wisdom"]), default="")
        sources.append(f"本地运营智慧库，更新时间：{latest or '未知'}")
    if context["merchants"]:
        latest = max((m.get("updated_at") or "" for m in context["merchants"]), default="")
        sources.append(f"本地商家资源库，更新时间：{latest or '未知'}")
    if context.get("raw_rows") or context.get("raw_files"):
        sources.append("本地原始数据湖")
    if context.get("holidays"):
        sources.append("本地日历数据")
    return "（信息来源：" + "；".join(sources or ["本地数据库暂无匹配信息"]) + "）"


def ask_deepseek_guidance(message):
    cfg = read_deepseek_config()
    if not cfg["key"]:
        return ""
    with get_conn() as conn:
        countries = [r["country"] for r in conn.execute("SELECT DISTINCT country FROM raw_files WHERE country <> '' ORDER BY country")]
        months = [f"{r['country']}{r['year']}年{r['month']}月" for r in conn.execute("SELECT DISTINCT country, year, month FROM raw_files ORDER BY country, year DESC, month DESC LIMIT 18")]
    prompt = (
        f"用户问题：{message}\n"
        f"当前本地库已有国家：{'、'.join(countries) or '暂无'}\n"
        f"当前本地库已有月份样例：{'、'.join(months) or '暂无'}\n"
        "本地库没有直接匹配结果。请生成2-3个中文引导性反问，格式为①... ②... ③...，不要编造具体销售结论。"
    )
    payload = {
        "model": "deepseek-chat",
        "messages": [
            {"role": "system", "content": "你是跨境电商数据助手。只生成引导性反问，不回答具体销售结论。"},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.3,
    }
    req = urllib.request.Request(
        cfg["base"] + "/chat/completions",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {cfg['key']}"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            return data["choices"][0]["message"]["content"].strip()
    except (urllib.error.URLError, KeyError, json.JSONDecodeError, TimeoutError):
        return ""


def local_guidance(message):
    terms = [t for t in search_terms(message) if t not in {"德国", "西班牙", "美国", "商家", "公司", "数据", "高峰"}]
    with get_conn() as conn:
        countries = [r["country"] for r in conn.execute("SELECT DISTINCT country FROM raw_files WHERE country <> '' ORDER BY country")]
        country = detect_country(message) or (countries[0] if countries else "德国")
        country_months = [f"{r['year']}年{r['month']}月" for r in conn.execute("SELECT DISTINCT year, month FROM raw_files WHERE country=? ORDER BY year DESC, month DESC LIMIT 3", (country,))]
    topic = terms[0] if terms else "相关类目"
    options = [
        f"{country}已上传哪些月份和类目的数据？",
        f"{country}{topic}有没有相近类目的销售结论？",
        f"商家资源库里有哪些{country}{topic}相关商家？",
    ]
    if country_months:
        options[0] = f"{country}{country_months[0]}的销售情况是什么？"
    CHAT_CONTEXT["suggestions"] = {str(i + 1): option for i, option in enumerate(options)}
    CHAT_CONTEXT["suggestions"].update({"①": options[0], "②": options[1], "③": options[2]})
    return "目前知识库中没有直接匹配到这个问题。您是否想问：\n" + "\n".join(f"{symbol}{option}" for symbol, option in zip(["①", "②", "③"], options))


def category_topic_from_message(message):
    fragments = sorted(meaningful_chinese_fragments(message), key=len, reverse=True)
    ignored = {"德国", "西班牙", "美国", "类目", "销售", "高峰", "价格", "价格带", "备货"}
    tree_rows = load_category_tree_rows()
    for fragment in fragments:
        if fragment in ignored:
            continue
        clean_fragment = clean_text(fragment)
        if any(clean_fragment and any(clean_fragment in clean_text(alias) for alias in category_aliases(row)) for row in tree_rows):
            return fragment
    for term in search_terms(message):
        if term not in ignored and term != message and re.search(r"[\u4e00-\u9fff]", term):
            return term
    return ""


def missing_category_data_answer(message, context):
    if not re.search(r"什么时候|高峰|卖得好|热销|备货|价格带", str(message or "")):
        return ""
    country = context.get("country") or detect_country(message)
    topic = category_topic_from_message(message)
    if country and topic:
        return f"尚未上传{country}{topic}类目的数据，请先上传相关销售数据。"
    return ""


def resolve_guided_message(message):
    text = str(message or "").strip()
    suggestions = CHAT_CONTEXT.get("suggestions") or {}
    if text in suggestions:
        return suggestions[text]
    digit = {"一": "1", "二": "2", "三": "3"}.get(text)
    if digit and digit in suggestions:
        return suggestions[digit]
    return message


def local_chat_answer(message):
    resolved_message = resolve_guided_message(message)
    context = fuzzy_search_all(resolved_message)
    if not context_has_results(context):
        answer = missing_category_data_answer(resolved_message, context)
        if answer:
            with get_conn() as conn:
                conn.execute("INSERT INTO chat_logs(role, content, created_at) VALUES ('user', ?, ?)", (message, now_text()))
                conn.execute("INSERT INTO chat_logs(role, content, created_at) VALUES ('assistant', ?, ?)", (answer, now_text()))
            return {"answer": answer, "context": context}
        rounds = CHAT_CONTEXT.get("guidance_rounds", 0) + 1
        CHAT_CONTEXT["guidance_rounds"] = rounds
        if rounds > 2:
            answer = "本地知识库中暂未找到相关信息，建议你上传对应的销售数据或手动更新知识库。"
        else:
            answer = local_guidance(resolved_message)
        with get_conn() as conn:
            conn.execute("INSERT INTO chat_logs(role, content, created_at) VALUES ('user', ?, ?)", (message, now_text()))
            conn.execute("INSERT INTO chat_logs(role, content, created_at) VALUES ('assistant', ?, ?)", (answer, now_text()))
        return {"answer": answer, "context": context}
    context = priority_context(context)
    CHAT_CONTEXT["suggestions"] = {}
    CHAT_CONTEXT["guidance_rounds"] = 0
    CHAT_CONTEXT["last_context"] = context
    answer = local_answer(context, resolved_message)
    answer = f"{answer}\n\n{source_note(context)}"
    with get_conn() as conn:
        conn.execute("INSERT INTO chat_logs(role, content, created_at) VALUES ('user', ?, ?)", (message, now_text()))
        conn.execute("INSERT INTO chat_logs(role, content, created_at) VALUES ('assistant', ?, ?)", (answer, now_text()))
    return {"answer": answer, "context": context}
